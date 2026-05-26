import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from tags import COLUMNS

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFD966")
HEADER_FONT = Font(bold=True)


def backup(input_path: str) -> str:
    """Copy input file to a timestamped backup. Returns backup path."""
    src = Path(input_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = src.parent / f"{src.stem}_backup_{timestamp}{src.suffix}"
    shutil.copy2(src, dest)
    return str(dest)


def load(path: str) -> pd.DataFrame:
    """Load first sheet of an xlsx file into a DataFrame."""
    return pd.read_excel(path, dtype=str)


def save(
    df: pd.DataFrame,
    changes: list[dict],
    yellow_cells: list[tuple[int, str]],
    output_path: str,
) -> None:
    """
    Write cleaned DataFrame to xlsx with:
      - Sheet 1 "Data":    cleaned data + yellow highlights
      - Sheet 2 "Deleted Data": cells where real content was erased
    """
    # Write base data via pandas (creates the file)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

        # Build change log DataFrame
        if changes:
            changes_df = pd.DataFrame(changes, columns=[
                "row", "company_name", "field", "old_value", "new_value", "highlighted"
            ])
            changes_df.columns = ["Row", "Company Name", "Field", "Old Value", "New Value", "Highlighted"]
            changes_df.insert(0, "#", range(1, len(changes_df) + 1))
        else:
            changes_df = pd.DataFrame(columns=["#", "Row", "Company Name", "Field", "Old Value", "New Value", "Highlighted"])

        changes_df.to_excel(writer, sheet_name="Deleted Data", index=False)

    # Re-open with openpyxl to apply formatting
    wb = load_workbook(output_path)
    _format_data_sheet(wb["Data"], df, yellow_cells)
    _format_changes_sheet(wb["Deleted Data"])
    wb.save(output_path)


def _format_data_sheet(ws, df: pd.DataFrame, yellow_cells: list[tuple[int, str]]) -> None:
    # Bold header row
    for cell in ws[1]:
        cell.font = HEADER_FONT

    # Build column index map: column_name → openpyxl column number (1-based)
    col_index = {col: i + 1 for i, col in enumerate(COLUMNS)}

    # Apply yellow fill (row offset: 1 header + 1-based → data row i has xlsx row i+2)
    yellow_set = {(int(row_idx) + 2, col_name) for row_idx, col_name in yellow_cells}
    for xlsx_row, col_name in yellow_set:
        col_num = col_index.get(col_name)
        if col_num:
            ws.cell(row=xlsx_row, column=col_num).fill = YELLOW_FILL

    # Auto-fit column widths (approximate)
    for col_num, col_name in enumerate(COLUMNS, start=1):
        max_len = max(
            len(col_name),
            df[col_name].astype(str).map(len).max() if col_name in df.columns else 0,
        )
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 50)


def _format_changes_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
    # Auto-fit
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
